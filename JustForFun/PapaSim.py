import math
import openpyxl

i4 = -131600
i5 = 25400
out = []
for i1 in range(0, 300, 50): #R1
    for i2 in range(0, 50, 10): #R2
        for i3 in range(0, 180, 15):
            p1 = [i4, i1, 0] #mikrometer


            x = round(i2 * math.cos(i3 * math.pi / 180), 8)
            y = round(i2 * math.sin(i3 * math.pi / 180), 8)

            p2 = [i5, x, y]

            pM = [0,0,0]

            x1 = p1
            x2 = [p1[0]-p2[0], p1[1]-p2[1], p1[2]-p2[2]] # v1 + lambda * v2

            # d = abs((pM - x1) x x2) / abs(x2)

            t1 = [pM[0]-x1[0], pM[1]-x1[1], pM[2]-x1[2]] #pM - x1
            t2 = [t1[1]*x2[2] - t1[2]*x2[1], t1[2]*x2[0] - t1[0]*x2[2], t1[0]*x2[1] - t1[1]*x2[0]]#t1 x x2
            t3 = math.sqrt(t2[0]*t2[0] + t2[1]*t2[1] + t2[2]*t2[2])
            t4 = math.sqrt(x2[0]*x2[0] + x2[1]*x2[1] + x2[2]*x2[2])
            t5 = t3/t4

            print(t5/1000)
            if t5 < 15 and t5 != 0:
                out.append([i1, i2, i3, t5])
            print("-"*25)
print("-"*25)
print("-"*25)
print("-"*25)
print("Format: [R1, R2, Grad, Abstand zum Mittelpunkt]")
for x in out:
    print(x)

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.cell(row=1, column=1, value="R1")
sheet.cell(row=1, column=2, value="R2")
sheet.cell(row=1, column=3, value="Grad")
sheet.cell(row=1, column=4, value="Abstand zum Mittelpunkt")

for row_idx, array in enumerate(out, start=2):
    for col_idx, value in enumerate(array, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)

# Speichere die Excel-Datei
workbook.save("output.xlsx")
